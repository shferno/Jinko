import sys
from PyQt5.QtWidgets import QApplication, QWidget
from PyQt5.QtCore import QCoreApplication, Qt, QEvent
from templates.weight import Ui_WeightRatioCalculator
import datetime
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, PatternFill, Protection, Side
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils.cell import get_column_letter
import os

path = "" # excel we want to store
class MyMainForm(QWidget, Ui_WeightRatioCalculator):
    def __init__(self):
        super(MyMainForm, self).__init__()
        self.setupUi(self)
        self.retranslateUi(self)
        self.CAV.setFocus()
        self.Submit.clicked.connect(self.cal_rat)
        self.Exit.clicked.connect(QCoreApplication.instance().quit)
        self.Change.long_pressed.connect(self.chg_MFV)
        self.CAV.setMaxLength(6)
        self.CBV.setMaxLength(6)
        self.TAV.setMaxLength(6)
        self.TBV.setMaxLength(6)
        self.MaN = "HUITIAN"
        self.MFV.setText(self.MaN)
        self.MFV.setFocusPolicy(Qt.NoFocus)
        self.count = [[0, 0], [0, 0], [0, 0], [0, 0]]
        QWidget.setTabOrder(self.CAV, self.CBV)
        QWidget.setTabOrder(self.CBV, self.TAV)
        QWidget.setTabOrder(self.TAV, self.TBV)
        QWidget.setTabOrder(self.TBV, self.MFV)
        QWidget.setTabOrder(self.MFV, self.INSPV)
        QWidget.setTabOrder(self.INSPV, self.Submit)
        QWidget.setTabOrder(self.Submit, self.Exit)
        self.RV.setFocusPolicy(Qt.NoFocus)
        self.comboBox.setFocusPolicy(Qt.NoFocus)
    def cal_rat(self):
        ca = self.CAV.text()
        cb = self.CBV.text()
        ta = self.TAV.text()
        tb = self.TBV.text()
        mf = self.MFV.text()
        ins = self.INSPV.text()
        comb = self.comboBox.currentText()
        raw_time = datetime.datetime.now().strftime("%Y/%m/%d/%H/%M/%S")
        time = raw_time.split("/")
        try:
            if not (ca.replace(".", "").isnumeric() and cb.replace(".", "").isnumeric and ta.replace(".", "").isnumeric() and tb.replace(".", "").isnumeric()):
                self.RV.append("Not a valid input weight, please check again. \n")
                self.clear_form()
            elif (float(ta) < float(ca) or float(tb) < float(cb) or float(tb) - float(cb) > float(ta) - float(ca)):
                self.RV.append("Wrong data entered, please do it again. \n")
                self.clear_form()
            else:
                ratio = round((float(ta) - float(ca)) / (float(tb) - float(cb)), 2)
                if not ins:
                    self.RV.append("Name required, please input the name.\n")
                    QWidget.setFocus(self.INSPV)
                else:
                    result = self.get_result(time, ratio, comb)
                    self.RV.append(f"{time[1]}/{time[2]}/{time[0]} {time[3]}:{time[4]}:{time[5]}\n"
                                   f"{comb}: the mixture ratio of glue is  *** {ratio} *** \n"
                                   f"Result: {result}\n")
                    title = pd.DataFrame(
                        columns=["Line No.", "Date", "Time", "Silicon Gel Filling Manufacturer", "Cup Weight A", "Total Weight A",
                                 "Silicone Gel Weight A", "Cup Weight B", "Total Weight B", "Silicone Gel Weight B", "Ratio",
                                 "Whether qualified", "Inspection Personnel"]
                    )
                    df = pd.DataFrame(
                        data=[[f"{comb}", f"{time[1]}/{time[2]}/{time[0]}", f"{time[3]}:{time[4]}:{time[5]}", mf, ca, ta,
                               float(ta) - float(ca), cb, tb, float(tb) - float(cb), ratio,
                               "Y" if ratio >= 5.5 and ratio <= 6.0 else "N", ins]],
                        columns=["Line No.", "Date", "Time", "Silicon Gel Filling Manufacturer", "Cup Weight A", "Total Weight A",
                                 "Silicone Gel Weight A", "Cup Weight B", "Total Weight B", "Silicone Gel Weight B", "Ratio",
                                 "Whether qualified", "Inspection Personnel"]
                    )

##################################################
#TBD:
                    if os.path.exists(os.getcwd() + r"\potting_glue_record.xlsx"):
                        # ew = StyleFrame.ExcelWriter(os.getcwd() + r"\potting_glue_record.xlsx", mode = "a", if_sheet_exists = "overlay")
                        # sf = StyleFrame(df)
                        # sf.to_excel(excel_writer = ew)
                        # ew.close()
                        # wt = pd.ExcelWriter(os.getcwd() + r"\potting_glue_record.xlsx")
                        # data = pd.read_excel(os.getcwd() + r"\potting_glue_record.xlsx", sheet_name = f"{comb}")
                        # data = pd.concat([data, df])
                        # data.to_excel(wt, index = None, sheet_name = f"{comb}")
                        # wt.close()
                        wb = openpyxl.load_workbook(os.getcwd() + r"\potting_glue_record.xlsx")
                        ws = wb[comb]
                        for row in dataframe_to_rows(df, index = False, header = False):
                            ws.append(row)
                        col_letter = []
                        for col in range(1, df.shape[1] + 1):
                            col_letter.append(get_column_letter(col))
                        for l in col_letter:
                            col_dim = ws.column_dimensions[l]
                            cont_wid = max(len(str(cell.value)) for cell in ws[l])
                            new_wid = (cont_wid + 2)
                            col_dim.width = new_wid
                        wb.save(os.getcwd() + r"\potting_glue_record.xlsx")
                    else:
                        # excel_writer = StyleFrame.ExcelWriter(os.getcwd() + r"\potting_glue_record.xlsx", mode = "w")
                        # sf = StyleFrame(df)
                        # sf.to_excel(
                        #     excel_writer = excel_writer,
                        #     best_fit = ["Date", "Time", "Silicon Gel Filling Manufacturer", "Cup Weight A1", "Total Weight A2",
                        #          "Silicone Gel Weight A", "Cup Weight B1", "Total Weight B2", "Silicone Gel Weight", "Ration",
                        #          "Whether qualified", "Inspection Personnel"],
                        #     # columns_and_rows_to_freeze = "L2",
                        #     # row_to_add_filters = 0,
                        # )
                        # excel_writer.close()
                        # df.to_excel(os.getcwd() + r"\potting_glue_record.xlsx", index = None, sheet_name = f"{comb}")
                        # wb = openpyxl.load_workbook(os.getcwd() + r"\potting_glue_record.xlsx")
                        # wb.create_sheet(title = f"{'Line 2' if comb == 'Line 1' else 'Line 1'}")
                        # wb.save(os.getcwd() + r"\potting_glue_record.xlsx")
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Line 1"
                        wb.create_sheet("Line 2")
                        for sn in wb.sheetnames:
                            ws = wb[sn]
                            for row in dataframe_to_rows(title, index = False, header = True):
                                ws.append(row)
                        ws = wb[comb]
                        for row in dataframe_to_rows(df, index = False, header = False):
                            ws.append(row)
                        col_letter = []
                        for col in range(1, df.shape[1] + 1):
                            col_letter.append(get_column_letter(col))
                        for l in col_letter:
                            col_dim = ws.column_dimensions[l]
                            cont_wid = max(len(str(cell.value)) for cell in ws[l])
                            new_wid = (cont_wid + 2)
                            col_dim.width = new_wid
                        wb.save(os.getcwd() + r"\potting_glue_record.xlsx")
####################################
                    self.clear_form()
        except Exception as e:
            print(e)
            self.RV.append("Wrong data entered, please do it again\n")
            self.clear_form()
            QWidget.setFocus(self.CAV)




    def clear_form(self):
        self.CAV.clear()
        self.CBV.clear()
        self.TAV.clear()
        self.TBV.clear()
        # self.MFV.clear()
        self.MFV.setText(self.MaN)
        self.INSPV.clear()
        QWidget.setFocus(self.CAV)
    def chg_MFV(self):
        QWidget.setFocus(self.MFV)


    def keyPressEvent(self, evt):
        QWidget.keyPressEvent(self, evt)
        if (evt.key() == Qt.Key_Enter or evt.key() == Qt.Key_Return):
            if QWidget.focusWidget(self).text():
                if QWidget.focusWidget(self) == self.INSPV:
                    self.cal_rat()
                    QWidget.setFocus(self.CAV)
                elif QWidget.focusWidget(self) == self.MFV:
                    self.MaN = self.MFV.text()
                    QWidget.clearFocus(self.MFV)
                    QWidget.setFocus(self.CAV)
                else:
                    QWidget.focusNextChild(self)

    def get_result(self, time, ratio, line):
        d_l = {"Line 1": [0, 1], "Line 2": [2, 3]}
        if float(time[3]) >= 6 and float(time[3]) < 18:
            if self.count[d_l[line][1]] != [0, 0]:
                self.count[d_l[line][1]] = [0, 0]
            if ratio >= 5.5 and ratio <= 6.0:
                self.count[d_l[line][0]][0] += 1
                result = "OK, "
                if self.count[d_l[line][0]][0] < 2:
                    result += "Still need do test again."
                else:
                    result += "Great, No need to test again."
            else:
                self.count[d_l[line][0]][1] += 1
                result = "NG, "
                if ratio < 5.5:
                    result += "lower ratio, "
                else:
                    result += "higher ratio, "
                if self.count[d_l[line][0]][1] < 2:
                    result += "Do the test again."
                else:
                    result += "Call Equipment and Process to fix!!!"
        else:
            if self.count[d_l[line][0]]!= [0, 0]:
                self.count[d_l[line][0]] = [0, 0]
            if ratio >= 5.5 and ratio <= 6.0:
                self.count[d_l[line][1]][0] += 1
                result = "OK"
                if self.count[d_l[line][1]][0] < 2:
                    result += "Still need do test again."
                else:
                    result += "Great, No need to test again."
            else:
                self.count[d_l[line][1]][1] += 1
                result = "NG, "
                if ratio < 5.5:
                    result += "lower ratio, "
                else:
                    result += "higher ratio, "
                if self.count[d_l[line][1]][1] < 2:
                    result += "Do the test again."
                else:
                    result += "Call Equipment and Process to fix!!!"
        return result


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWin = MyMainForm()
    myWin.show()
    sys.exit(app.exec())